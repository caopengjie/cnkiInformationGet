#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Created by  CjEff on 2018/10/23

import re
import time
import xlrd
from openpyxl import load_workbook
from selenium import webdriver


# 信息
inforGet ='Names.xlsx'
inforSave = 'Information.xlsx'

# Chromedriver path
ChromedriverPath = 'C:/Python37/Scripts/chromedriver'

class CnkiInforGet(object):
    def __init__(self):
        self.driver = webdriver.Chrome(ChromedriverPath)
        self.driver.set_window_size(1400, 800)
        self.inforInit()
        self.getData()

    def getData(self):
        yearFrom = '2007'
        yearTo = '2016'
        # 创建存储的表格
        wb = load_workbook(inforSave) # 在内存中创建一个workbook对象
        ws = wb.active
        # 创建存储的title
        ws.cell(row=1, column=1).value = '题名'
        ws.cell(row=1, column=2).value = '作者'
        ws.cell(row=1, column=3).value = '来源'
        ws.cell(row=1, column=4).value = '发表时间'
        ws.cell(row=1, column=5).value = '数据库'
        ws.cell(row=1, column=6).value = '被引'
        ws.cell(row=1, column=7).value = '下载'
        ws.cell(row=1, column=8).value = '摘要'
        ws.cell(row=1, column=9).value = '页数'
        ws.cell(row=1, column=10).value = '关键词'

        year_from = self.driver.find_element_by_id('year_from')
        year_to = self.driver.find_element_by_id('year_to')
        textInput = self.driver.find_element_by_id('magazine_value1')
        btnsearch = self.driver.find_element_by_id('btnSearch')
        num = 1
        for magzines in self.searchNames:
            year_from.send_keys(yearFrom)
            year_to.send_keys(yearTo)
            textInput.send_keys(magzines)
            btnsearch.click()
            textInput.clear()
            time.sleep(1)
            self.driver.switch_to.frame("iframeResult")
            countRecord = "".join(re.findall("\d+", self.driver.find_element_by_xpath('//div[@class="pagerTitleCell"]').text))
            pages = int(int(countRecord)/20)

            #GetData from pages
            for page in range(pages):
                #  按行查询表格的数据，再按照列拆分数据
                table_loc = self.driver.find_element_by_xpath('//table[@class="GridTableContent"]')
                table_tr_list = self.driver.find_elements_by_tag_name('tr')
                i = 0
                # 遍历每一个tr，并将其中每一个td的数据查询出来
                for tr in table_tr_list:
                    i = i + 1
                    if i >= 8 and i <=27:
                        table_td_list = tr.find_elements_by_tag_name('td')
                        row_list = []
                        # 选择固定的行去遍历数据,前面以及后面几个行
                        for td in table_td_list:
                            row_list.append(td.text)
                        if row_list[2]!='':
                            num = num + 1
                            # 将数据写入到excel当中
                            print(num)
                            ws.cell(row=num, column=1).value = row_list[1]
                            ws.cell(row=num, column=2).value = row_list[2]
                            ws.cell(row=num, column=3).value = row_list[3]
                            ws.cell(row=num, column=4).value = row_list[4]
                            ws.cell(row=num, column=5).value = '期刊'
                            ws.cell(row=num, column=6).value = row_list[5]
                            ws.cell(row=num, column=7).value = row_list[6]

                            # 获取摘要、页数、关键词
                            self.driver.find_element_by_link_text(row_list[1]).click()
                            store_handle = self.driver.current_window_handle
                            self.driver.switch_to.window(self.driver.window_handles[-1])
                            print (self.driver.window_handles)
                            abstract = self.driver.find_element_by_xpath('//*[@id="ChDivSummary"]').text
                            paperNum = self.driver.find_element_by_xpath('//*[@id="mainArea"]/div[3]/div[3]/div[1]/div[4]/div[1]/div[1]/span[3]/b').text
                            keyWords = self.driver.find_element_by_xpath('//*[@id="mainArea"]/div[3]/div[3]/div[1]/p[3]').text
                            keywords_handle = keyWords[4:]
                            # 将数据写入到excel当中
                            ws.cell(row=num, column=8).value = abstract
                            ws.cell(row=num, column=9).value = paperNum
                            ws.cell(row=num, column=10).value = keywords_handle
                            self.driver.close()
                            self.driver.switch_to.window(store_handle)
                            self.driver.switch_to.frame("iframeResult")
                            time.sleep(1)
                time.sleep(2)
                # 请求下一页
                self.driver.find_element_by_link_text(u'下一页').click()
            self.driver.switch_to.default_content()
            time.sleep(5)
            # 保存数据文件
            wb.save(inforSave)



    def inforInit(self):
        # get names from xlsx
        self.searchNames = self.GetSearchNames()
        # init search page
        self.initPage()

    def initPage(self):
        self.driver.get('http://kns.cnki.net/kns/brief/result.aspx?dbprefix=CJFQ')
        time.sleep(1)

    def GetSearchNames(self):
        workbook = xlrd.open_workbook(inforGet)
        booksheet = workbook.sheet_by_index(0)
        search_iters = booksheet.col_values(0)
        return search_iters

if __name__ == '__main__':
    test = CnkiInforGet()
